import csv
from datetime import date
from openpyxl import load_workbook, Workbook

from importer import *


class FileImporter:

    def __init__(self, filePath, defaultHeaders=None, defaultDataTypes=None, rowDataType=ROW_TYPE_DICT,
                 noneStrings=DEFAULT_NONE_STRINGS, rowLimit=None, rowFilterFn=None):
        """A class used to read, scrub, analyze, and write data from/to Excel and CSV

        :param str filePath: Name of the file to read
        :param dict defaultHeaders: A key/value pair of header in raw file and the new name of header
        :param dict defaultDataTypes: A key/value pair of header and the function or data type to convert each value to
        :param str rowDataType: A string indicating the data type that each row should be processed as
        :param iterable noneStrings: String values that should be converted to None
        :param int rowLimit: The number of rows to get from the import file
        :param function rowFilterFn: A function which returns True if a row should be included
        """
        self.defaultHeaders = defaultHeaders
        self.defaultDataTypes = defaultDataTypes
        self.rowDataType = rowDataType
        self.noneStrings = noneStrings

        self.fileReader = self.getFileReader(filePath)
        self.headers = self.processHeaders(list(next(self.fileReader)))
        print(f'Headers: {self.headers}')
        self.data = self.getData(rowLimit, rowFilterFn)

    ##### READ DATA #########
    def printRows(self, numRows):
        for idx, row in enumerate(self.data):
            print(row)
            if idx == numRows - 1:
                break

    def getData(self, rowLimit, rowFilterFn):
        """ Get all row data from the file import and return a processed list of records
        :return: list
        """
        data = []
        for idx, row in enumerate(self.fileReader):
            if rowLimit and len(data) == rowLimit:
                break
            processedRow = self.processRow(row)
            if idx == 0:
                headersAndVals = zip(self.headers, processedRow) if self.rowDataType != ROW_TYPE_DICT else processedRow.items()
                for header, val in headersAndVals:
                    print(f'{header} type is {type(val)}')
            if rowFilterFn and not rowFilterFn(processedRow):
                continue
            data.append(processedRow)

        # CSV file must be closed manually
        if self.fileType == FILE_TYPE_CSV:
            self.file.close()

        return data

    def processRow(self, row):
        """ Process the row into the data type specified by FileImporter
        :return: list|tuple|dict
        """
        processedRow = []
        for header, val in zip(self.headers, row):
            defaultDataType = self.defaultDataTypes.get(header) if self.defaultDataTypes else None
            processedRow.append(self.processValue(val, defaultDataType))
        if self.rowDataType == ROW_TYPE_LIST:
            return processedRow
        if self.rowDataType == ROW_TYPE_DICT:
            return {header: val for header, val in zip(self.headers, processedRow)}
        if self.rowDataType == ROW_TYPE_TUPLE:
            return tuple(processedRow)

    def processValue(self, val, defaultDataType):
        """ Get a processed value. Handles None, null strings, and function conversions
        :param func|class defaultDataType: A function or class which converts the value into the appropriate format
        """
        if val is None:
            return val
        if isinstance(val, str):
            val = val.strip()
            if val.lower() in self.noneStrings:
                return None
        if defaultDataType:
            val = defaultDataType(val)
        return val

    def getFileReader(self, filePath):
        """ Get an iterator used to get each row in a CSV or XLSX file
        """
        if f'.{FILE_TYPE_CSV}' in filePath:
            self.file = open(filePath)
            self.fileType = FILE_TYPE_CSV
            return csv.reader(self.file)
        elif f'.{FILE_TYPE_XLS}' in filePath:
            self.file = load_workbook(filePath, data_only=True)
            worksheet = self.file.active
            self.fileType = FILE_TYPE_XLS
            return worksheet.iter_rows(values_only=True)
        else:
            raise(ValueError('Unsupported file type'))

    def processHeaders(self, headers):
        """ Convert headers to formatted names
        """
        if not self.defaultHeaders:
            return headers
        newHeaders = []
        for idx, header in enumerate(headers):
            headerByIdx = self.defaultHeaders.get(idx)
            headerByVal = self.defaultHeaders.get(header)
            if headerByIdx:
                newHeaders.append(headerByIdx)
            elif headerByVal:
                newHeaders.append(headerByVal)
            else:
                newHeaders.append(header)
        return newHeaders

    ##### WRITE DATA #########
    def writeCsvFile(self, fileName, data=None, headers=None):
        """ Write data to a new CSV file
        :param str fileName: The name of the file to write to. The file extension should not be included.
        :param list data: If provided, will be used instead of the FileImporter's internal data property
        :param list headers: If provided, will be used instead of the FileImporter's internal header property
        """
        dataToWrite = data or self.data
        headersToWrite = headers or self.headers

        with open(f'{DATA_FILE_OUTPUT_PATH}{fileName}_{date.today()}.csv', 'w') as csvFile:
            csvWriter = csv.writer(csvFile)
            csvWriter.writerow(headersToWrite)
            for row in dataToWrite:
                csvWriter.writerow(self.formatRow(row))

    def writeExcelFile(self, fileName, sheetsConfig=None):
        """ Write data to a new XLSX file and return the workbook
        :param str fileName: The name of the file to write to. The file extension should not be included.
        :param list sheetsConfig: [{'title': , 'data': , 'headers':},...] If provided, will be used instead
        of the FileImporter's internal data property
        """
        if not sheetsConfig:
            return

        workbook = Workbook()

        for idx, config in enumerate(sheetsConfig):
            # Create new worksheet
            if idx == 0:
                worksheet = workbook.active
            else:
                worksheet = workbook.create_sheet()
            worksheet.title = config.get('title', None)

            # Add data to sheet
            dataToWrite = config.get('data', None) or self.data
            headersToWrite = config.get('headers', None) or self.headers
            worksheet.append(headersToWrite)
            for row in dataToWrite:
                worksheet.append(self.formatRow(row))

        # Save and return workbook
        workbook.save(f'{DATA_FILE_OUTPUT_PATH}{fileName}_{date.today()}.xlsx')
        return workbook

    def formatRow(self, row):
        """ Process row to make sure it is an appropriate data format to write to a CSV or XSLX file.
        """
        if isinstance(row, (tuple, list)):
            return row
        if isinstance(row, dict):
            return list(row.values())
        raise ValueError('Must use a row data type of tuple, list, or dict')

    ##### ANALYZE DATA #########
    # [(customerKey, (age, <function to group>), (sex, isMarried)]
    # {('customerKey'): {'a-11243d': [<record1>, <record2>]}, (sex, isMarried): {}}
    def getGroupData(self, groupings, filterFn=None, data=None, headers=None, isFlat=False):
        """Get group segments for each distinct key generated by the column keys for each grouping
        The return value will contain the group segment key as the key and the list of records that have
        that group segment as the key value.
        :param list groupings: A list of tuples. Each tuple should contain one or more column keys. Optionally
        the tuple value camn be another tuple with the first value expressing the column key and the second
        value expressing a function which will be used to format/group the raw value (e.g. a function which
        takes an age number and formats it into an age category).
        :param function filterFn: A function that returns a boolean value, true indicating that the record value should
        be included
        Example:
            input: [('occupation', ('age', <function to group>)), ('sex', 'isMarried')]
            output: {
                ('occupation', 'age'): {
                    ('lawyer', '30-40'): [<record1>, <record2>, ...],
                    ('lawyer', '40-50'): [<record1>, <record2>, ...],
                    ...
                    }
                ('sex', 'isMarried'): {...}
            }
        :param list data: Optionally pass in data records. If not passed in, self.data will be used.
        :param list headers: Optionall pass in headers. If not passed in, self.headers will be used.
        If data is passed in and records are not dictionaries, you must also pass in headers.
        :param bool isFlat: If true, only values will be returned
        """
        data = data or self.data
        headers = headers or self.headers
        dataGroups = {}
        for group in groupings:
            dataGroups[self.getGroupColumnKey(group)] = {}

        for group in groupings:
            currentGroup = dataGroups[self.getGroupColumnKey(group)]
            for record in data:
                if not filterFn or filterFn(record):
                    groupKey = self.getGroupKey(record, group, headers)
                    if groupKey in currentGroup:
                        currentGroup[groupKey].append(record)
                    else:
                        currentGroup[groupKey] = [record]

        return dataGroups if not isFlat else list(dataGroups.values())

    def setGroupData(self, groupings):
        self.dataGroups = self.getGroupData(groupings)

    def getGroupColumnKey(self, group):
        groupingColumnKeys = []
        for column in group:
            if isinstance(column, str):
                groupingColumnKeys.append(column)
            elif isinstance(column, tuple):
                groupingColumnKeys.append(column[0])
            else:
                raise ValueError('Column must be string or tuple')
        return tuple(groupingColumnKeys)

    def getGroupKey(self, record, group, headers):
        """Get the record values corresponding with the group column keys. for example, if the group
        is ('occupation', 'isMarried'), the return value would be something lict ('doctor', 0).
        :param record: An iterable that contains data for a single record
        :param tuple group: Tuple of column keys and/or tuple with (<columnKey>, <function>)
        :param list headers: Header values for each column in record.
        :return: Tuple of values for each column key
        """
        if not isinstance(record, dict):
            record = {header: val for header, val in zip(headers, record)}

        groupKey = []
        for column in group:
            if isinstance(column, tuple):
                columnKey, groupingFunc = column
                val = groupingFunc(record[columnKey])
            else:
                val = record[column]
            groupKey.append(val)
        return tuple(groupKey)
